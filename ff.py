import cv2
import os
import numpy as np
from datetime import datetime
import openpyxl

def load_trainers():
    trainers = {}
    for user_dir in os.listdir(r'Datasets'):
        id = user_dir.replace('user', '')
        trainer_file = os.path.join(r'Datasets', user_dir, 'trainer.yml')
        if os.path.exists(trainer_file):
            recognizer = cv2.face.LBPHFaceRecognizer.create()
            recognizer.read(trainer_file)
            trainers[id] = recognizer
    return trainers

def recognize(trainers):
    tn = datetime.now()
    t = tn.strftime('%H:%M:%S')
    d = tn.strftime('%d-%m-%Y')

    video = cv2.VideoCapture(0)
    # address='https://192.168.137.68:8080/video'
    # video.open(address)
    facedetect = cv2.CascadeClassifier(r'haarcascade_frontalface_default.xml')

    while True:
        ret, frame = video.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = facedetect.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            for id, recognizer in trainers.items():
                serialno, conf = recognizer.predict(gray[y:y+h, x:x+w])
                if conf < 50:
                    cv2.putText(frame, id, (x, y-40), cv2.FONT_HERSHEY_SIMPLEX, 1, (50, 50, 225), 2)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (50, 50, 255), 2)
                    id=int(id)
                    tn = datetime.now()
                    t = tn.strftime('%H:%M:%S')
                    d = tn.strftime('%d-%m-%Y')
                    if os.path.exists('ecount.txt'):
                        with open('ecount.txt','r')as exc:
                            ec=exc.read()
                            ec=int(ec)
                    else:  
                        ec=2
                        ec=int(ec)
                    path=r'Saved_Plates\1StudentDatabase.xlsx'
                    wb = openpyxl.load_workbook(path)
                    ws = wb.active
                    ws.cell(row=ec,column=2,value=t)
                    ws.cell(row=ec,column=1,value=d)
                    ws.cell(row=ec ,column=3,value = id)
                    with open(rf'Datasets\{id}_name.txt','r') as f:
                        n=f.read()
                    ws.cell(row=ec ,column=4,value = n)
                    ec+=1
                    with open('ecount.txt','w') as exc:
                        exc.write(str(ec))
                    wb.save(path)
                    wb.close()
                else:
                    cv2.putText(frame, '', (x, y-40), cv2.FONT_HERSHEY_SIMPLEX, 1, (50, 50, 225), 2)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (50, 50, 255), 1)

        cv2.imshow('Face', frame)
        k = cv2.waitKey(1)
        if k == 27 or k == 'q':
            break

    cv2.destroyAllWindows()

def training(id):
    video = cv2.VideoCapture(0)
    facedetect = cv2.CascadeClassifier(r'haarcascade_frontalface_default.xml')

    count = 0
    while True:
        ret, frame = video.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = facedetect.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x+w, y+h), (50, 50, 255), 1)
            count += 1
            if not os.path.exists('datasets/user' + id):
                os.mkdir('datasets/user' + id)
            cv2.imwrite('datasets/user{}/user.{}.{}.jpg'.format(id, id, count), gray[y:y+h, x:x+w])

        cv2.imshow('Face', frame)
        k = cv2.waitKey(1)
        if count>200:  
            break

    cv2.destroyAllWindows()
    print('DataSet Collection Done.......')

    recognizer = cv2.face.LBPHFaceRecognizer.create()
    id = int(id)
    path = r'Datasets\user{}'.format(id)

    def getImageID(path):
        imagepath = [os.path.join(path, f) for f in os.listdir(path)]
        faces = []
        ids = []
        for imagepaths in imagepath:
            faceimage = cv2.imread(imagepaths, cv2.IMREAD_GRAYSCALE)
            faceNP = np.array(faceimage, 'uint8')
            faces.append(faceNP)
            ids.append(id)
            cv2.imshow('Training...', faceNP)
            cv2.waitKey(1)
        return ids, faces

    IDs, facedata = getImageID(path)
    recognizer.train(facedata, np.array(IDs))
    recognizer.write(r'Datasets\user{}\trainer.yml'.format(id))
    cv2.destroyAllWindows()
    print('Training Completed for ID:', id)

