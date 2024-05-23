import json             
import requests
import openpyxl
from datetime import datetime

tn=datetime.now()
t=tn.strftime('%H:%M:%S')
d=tn.strftime('%d-%m-%Y')


def Upload(Licence,count):
    Number  = str(Licence)
    c = int(count)
    c += 1
    url = 'https://www.carinfo.app/_next/data/m3jXnTcG9Plj9OMSj4dEF/rc-details/' + Number + '.json?rc=' + Number

    response = requests.get(url).json()
    

    Reg_Address = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['messages'][1]['subtitle']
    City = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['messages'][2]['subtitle']
    State = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['messages'][3]['subtitle']
    Name = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['message']['title']
    Carname = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['message']['subtitle']
    path=r'D:\DRP Project\Saved_Plates\Database.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.cell(row=c,column=1,value=t)
    ws.cell(row=c,column=2,value=d)
    ws.cell(row=c ,column=3,value = Number)
    ws.cell(row=c ,column=4,value = Reg_Address)
    ws.cell(row=c ,column=5,value = City)
    ws.cell(row=c ,column=6,value = State)
    ws.cell(row=c ,column=7,value = Name)
    ws.cell(row=c ,column=8,value = Carname)

    wb.save(path)
    wb.close()
    
    

i = 'WB06 F 5977'
Upload(i,9)

if'__name__'=='__main__':
    Upload()


