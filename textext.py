def textext(link):
    import requests

    url = "https://api.apilayer.com/image_to_text/upload"


    with open(link, 'rb') as file:
        file_content = file.read()


    headers = {
        "apikey":#apikey
    }

    response = requests.post(url, headers=headers, data=file_content)
    status_code = response.status_code
    result = response.json()
    print(result)
    plno=result["all_text"]
    plno=str(plno)

    if "IND" in plno :
        plno = plno[4::]
    print(plno)

    import json             
    import requests
    import openpyxl
    from datetime import datetime

    tn=datetime.now()
    t=tn.strftime('%H:%M:%S')
    d=tn.strftime('%d-%m-%Y')

    with open('c.txt','r')as exc:
        ec=exc.read()
        ec=int(ec)
    try:
        def Upload(Licence,count):
            Number  = str(Licence)
            c = int(count)
            c += 1
            url = 'https://www.carinfo.app/_next/data/MuceL2U-BNrimc0ehV-rh/rc-details/' + Number + '.json?rc=' + Number

            response = requests.get(url).json()
            

            Reg_Address = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['messages'][1]['subtitle']
            City = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['messages'][2]['subtitle']
            State = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['messages'][3]['subtitle']
            Name = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['message']['title']
            Carname = response['pageProps']['rcDetailsResponse']['data']['webSections'][0]['message']['subtitle']
            path=r'Saved_Plates\Database.xlsx'
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ws.cell(row=c,column=1,value=t)
            ws.cell(row=c,column=2,value=d)
            ws.cell(row=c ,column=3,value = Number)
            ws.cell(row=c ,column=4,value = Reg_Address)
            ws.cell(row=c ,column=5,value = City)
            ws.cell(row=c ,column=6,value = State)
            ws.cell(row=c ,column=7,value = Name)
            ws.cell(row=c ,column=8,value = Carname)

            wb.save(path)
            wb.close()
    except KeyError:
        print('Issue with the backend')
    except requests.exceptions.RequestException as e:
            print(f"HTTP error occurred: {e}")
        

    i = plno
    Upload(i,ec)
    ec+=1
    with open('c.txt','w') as exc:
        exc.write(str(ec))

if __name__=='__main__':
    e=textext()



